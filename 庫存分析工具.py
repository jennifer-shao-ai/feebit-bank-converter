#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
飛比特 庫存訂貨分析工具
用途：讀取商品總表 CSV，過濾套組商品，標色警示，產出 Excel 訂貨分析表
執行：python3 庫存分析工具.py [CSV檔案路徑]
     若不指定路徑，預設讀取同目錄的「全產品總表.csv」
"""

import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 設定 ──────────────────────────────────────────────────────────────────────

DEFAULT_CSV = os.path.join(os.path.dirname(__file__), "全產品總表.csv")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "庫存訂貨分析.xlsx")

# 色碼定義
FILLS = {
    'red':    PatternFill(start_color='FF5252', end_color='FF5252', fill_type='solid'),   # 庫存為 0
    'blue':   PatternFill(start_color='64B5F6', end_color='64B5F6', fill_type='solid'),   # 低於安全庫存
    'orange': PatternFill(start_color='FFB74D', end_color='FFB74D', fill_type='solid'),   # 低於 20
    'yellow': PatternFill(start_color='FFF176', end_color='FFF176', fill_type='solid'),   # 低於 50
    'purple': PatternFill(start_color='CE93D8', end_color='CE93D8', fill_type='solid'),   # 無安全庫存但低於 10
    'header': PatternFill(start_color='37474F', end_color='37474F', fill_type='solid'),   # 標題列
    'legend_title': PatternFill(start_color='455A64', end_color='455A64', fill_type='solid'),
}

HEADER_FONT = Font(color='FFFFFF', bold=True, name='微軟正黑體', size=11)
NORMAL_FONT = Font(name='微軟正黑體', size=10)

# ── 核心函數 ──────────────────────────────────────────────────────────────────

def is_bundle(code: str) -> bool:
    """套組判斷：代號恰好 8 碼，第 3 個字元（index 2）為 '9'"""
    code = str(code).strip()
    return len(code) == 8 and code[2] == '9'


def clean_number(val) -> float | None:
    """清理數字欄位（處理 '87.'、'-10.' 等格式）"""
    if pd.isna(val):
        return None
    val = str(val).strip().rstrip('.')
    if val == '' or val == '-':
        return None
    try:
        return float(val)
    except ValueError:
        return None


def suggest_order(safety: float | None, qty: float | None) -> int | str:
    """
    建議訂購數量
    安全庫存 = 90 天用量，目標補到 180 天（2× 安全庫存）
    公式：max(0, 2 × safety − qty)，至少補到 safety
    """
    if safety is None or safety <= 0:
        return ''
    qty = qty if qty is not None else 0
    if qty < safety:
        suggested = max(int(safety * 2 - qty), int(safety))
        return suggested
    return 0


def get_fill_color(qty: float | None, safety: float | None) -> PatternFill | None:
    """依優先順序回傳對應的色碼"""
    if qty is None:
        qty = 0
    # 優先順序：紅 > 藍 > 橘 > 黃 > 紫
    if qty <= 0:
        return FILLS['red']
    if safety is not None and safety > 0 and qty < safety:
        return FILLS['blue']
    if qty < 20:
        return FILLS['orange']
    if qty < 50:
        return FILLS['yellow']
    if (safety is None or safety <= 0) and qty < 10:
        return FILLS['purple']
    return None


# ── 主流程 ────────────────────────────────────────────────────────────────────

def main(csv_path: str):
    print(f"📂 讀取：{csv_path}")

    # 讀取 CSV（支援 UTF-8 BOM 與一般 UTF-8）
    try:
        df = pd.read_csv(csv_path, encoding='utf-8-sig', dtype=str)
    except UnicodeDecodeError:
        df = pd.read_csv(csv_path, encoding='cp950', dtype=str)

    print(f"   原始資料：{len(df)} 筆")

    # 欄位對應（根據實際 CSV 欄位名稱）
    col_code    = df.columns[0]   # 商品代號
    col_name    = df.columns[1]   # 商品名稱
    col_safety  = df.columns[2]   # 安全庫存
    col_ean     = df.columns[3]   # EAN13碼
    col_status  = df.columns[4]   # 販售狀態
    col_status2 = df.columns[5]   # 狀態2
    col_loc     = df.columns[6]   # 庫位
    col_qty     = df.columns[7]   # 數量
    col_update  = df.columns[8]   # 最後更新日期
    col_create  = df.columns[9]   # 建立日期

    # 過濾套組商品
    mask_bundle = df[col_code].apply(is_bundle)
    df_filtered = df[~mask_bundle].copy()
    print(f"   過濾套組後：{len(df_filtered)} 筆（移除 {mask_bundle.sum()} 筆套組）")

    # 清理數字欄位
    df_filtered['_qty']    = df_filtered[col_qty].apply(clean_number)
    df_filtered['_safety'] = df_filtered[col_safety].apply(clean_number)

    # 計算建議訂購數量
    df_filtered['建議訂購數量'] = df_filtered.apply(
        lambda r: suggest_order(r['_safety'], r['_qty']), axis=1
    )
    df_filtered['實際訂購數量'] = ''

    # ── 建立 Excel ─────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "庫存訂貨分析"

    # 輸出欄位順序
    out_cols = [
        col_code, col_name, col_safety, col_ean,
        col_status, col_status2, col_loc, col_qty,
        col_update, col_create,
        '建議訂購數量', '實際訂購數量'
    ]

    # 標題列
    ws.append(out_cols)
    header_row = ws[1]
    for cell in header_row:
        cell.fill = FILLS['header']
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 32

    # 細邊框樣式
    thin = Side(style='thin', color='BDBDBD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 資料列
    for _, row in df_filtered.iterrows():
        values = [row.get(c, '') for c in out_cols]
        # 建議 / 實際訂購數量
        values[-2] = row['建議訂購數量']
        values[-1] = row['實際訂購數量']
        ws.append(values)

        row_num = ws.max_row
        qty    = row['_qty']
        safety = row['_safety']
        fill   = get_fill_color(qty, safety)

        for col_idx, cell in enumerate(ws[row_num], start=1):
            cell.font = NORMAL_FONT
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if fill:
                cell.fill = fill

    # 凍結標題列
    ws.freeze_panes = 'A2'

    # 自動欄寬
    col_widths = {
        1: 14,   # 商品代號
        2: 42,   # 商品名稱
        3: 12,   # 安全庫存
        4: 16,   # EAN13
        5: 12,   # 販售狀態
        6: 12,   # 狀態2
        7: 10,   # 庫位
        8: 10,   # 數量
        9: 18,   # 最後更新
        10: 14,  # 建立日期
        11: 14,  # 建議訂購數量
        12: 14,  # 實際訂購數量
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 色碼說明頁 ─────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("色碼說明")

    legend = [
        ("色碼說明", None, None),
        ("顏色", "條件", "建議動作"),
        ("紅色",    "庫存數量 = 0",              "立即補貨（最高優先）"),
        ("藍色",    "庫存低於安全庫存量",          "儘快補貨"),
        ("橘色",    "庫存低於 20（無安全庫存設定）", "視銷況考慮補貨"),
        ("黃色",    "庫存低於 50（無安全庫存設定）", "列入觀察"),
        ("紫色",    "未設安全庫存，但數量低於 10",  "確認是否需要補貨"),
        ("無顏色",  "庫存正常",                   "無需處理"),
    ]

    legend_fills = {
        "紅色": FILLS['red'],
        "藍色": FILLS['blue'],
        "橘色": FILLS['orange'],
        "黃色": FILLS['yellow'],
        "紫色": FILLS['purple'],
    }

    for i, (a, b, c) in enumerate(legend, start=1):
        row_data = [a, b, c] if b else [a]
        ws2.append(row_data)
        r = ws2[i]

        if i == 1:
            # 大標題
            r[0].fill = FILLS['legend_title']
            r[0].font = Font(color='FFFFFF', bold=True, name='微軟正黑體', size=13)
            r[0].alignment = Alignment(horizontal='center', vertical='center')
            ws2.merge_cells(f'A1:C1')
            ws2.row_dimensions[1].height = 30
        elif i == 2:
            # 表頭
            for cell in r:
                cell.fill = FILLS['header']
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            # 資料列
            fill = legend_fills.get(a)
            if fill:
                r[0].fill = fill
            for cell in r:
                cell.font = Font(name='微軟正黑體', size=10)
                cell.border = border
                cell.alignment = Alignment(vertical='center')

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 36
    ws2.column_dimensions['C'].width = 28

    # ── 統計摘要頁 ────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("庫存統計")

    total = len(df_filtered)
    qty_series   = df_filtered['_qty'].fillna(0)
    safe_series  = df_filtered['_safety']

    cnt_zero     = int((qty_series <= 0).sum())
    cnt_below_s  = int(((qty_series > 0) & (safe_series > 0) & (qty_series < safe_series)).sum())
    cnt_below_20 = int(((qty_series > 0) & (qty_series < 20) & ((safe_series.isna()) | (safe_series <= 0))).sum())
    cnt_below_50 = int(((qty_series >= 20) & (qty_series < 50) & ((safe_series.isna()) | (safe_series <= 0))).sum())
    cnt_purple   = int(((qty_series > 0) & (qty_series < 10) & ((safe_series.isna()) | (safe_series <= 0))).sum())
    cnt_normal   = total - cnt_zero - cnt_below_s - cnt_below_20 - cnt_below_50

    stats = [
        ("飛比特 庫存統計摘要",),
        ("項目", "數量", "佔比"),
        ("總商品數（不含套組）", total, "100%"),
        ("🔴 庫存為 0", cnt_zero,     f"{cnt_zero/total*100:.1f}%" if total else "0%"),
        ("🔵 低於安全庫存", cnt_below_s,  f"{cnt_below_s/total*100:.1f}%" if total else "0%"),
        ("🟠 低於 20（無安全庫存）", cnt_below_20, f"{cnt_below_20/total*100:.1f}%" if total else "0%"),
        ("🟡 低於 50（無安全庫存）", cnt_below_50, f"{cnt_below_50/total*100:.1f}%" if total else "0%"),
        ("🟣 未設安全庫存且低於 10", cnt_purple, f"{cnt_purple/total*100:.1f}%" if total else "0%"),
        ("✅ 庫存正常", cnt_normal, f"{cnt_normal/total*100:.1f}%" if total else "0%"),
    ]

    for i, row_data in enumerate(stats, start=1):
        ws3.append(list(row_data))
        r = ws3[i]
        if i == 1:
            r[0].fill = FILLS['legend_title']
            r[0].font = Font(color='FFFFFF', bold=True, name='微軟正黑體', size=13)
            r[0].alignment = Alignment(horizontal='center')
            ws3.merge_cells('A1:C1')
            ws3.row_dimensions[1].height = 30
        elif i == 2:
            for cell in r:
                cell.fill = FILLS['header']
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center')
        else:
            for cell in r:
                cell.font = Font(name='微軟正黑體', size=10)
                cell.border = border
                cell.alignment = Alignment(vertical='center')

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 12

    # ── 儲存 ──────────────────────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print(f"\n✅ 完成！輸出：{OUTPUT_FILE}")
    print(f"\n📊 庫存統計：")
    print(f"   總商品數（不含套組）：{total}")
    print(f"   🔴 庫存為 0      ：{cnt_zero}")
    print(f"   🔵 低於安全庫存  ：{cnt_below_s}")
    print(f"   🟠 低於 20       ：{cnt_below_20}")
    print(f"   🟡 低於 50       ：{cnt_below_50}")
    print(f"   🟣 未設且低於 10 ：{cnt_purple}")
    print(f"   ✅ 正常           ：{cnt_normal}")


if __name__ == '__main__':
    csv_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_CSV
    if not os.path.exists(csv_path):
        print(f"❌ 找不到檔案：{csv_path}")
        print(f"   用法：python3 庫存分析工具.py [CSV路徑]")
        print(f"   例如：python3 庫存分析工具.py ~/Downloads/全產品總表.csv")
        sys.exit(1)
    main(csv_path)
